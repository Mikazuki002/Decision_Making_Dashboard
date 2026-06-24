"""Views for the Office DTR system.

Key behaviors:
* List views always sort alphabetically (models.Meta.ordering + explicit
  .order_by() in the view).
* Search filters by name, employee id, or department.
* Editing any row triggers a full page re-render so the table updates
  automatically (auto-update on any change).
* Download endpoint streams the same filtered data as CSV or Excel.
"""
import csv
import hashlib
import json
from collections import defaultdict
from datetime import datetime, date, time, timedelta

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count, Avg, F
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.views.decorators.http import require_http_methods

from .models import Employee, Attendance, Department
from .forms import AttendanceForm, EmployeeForm
from .dashboard import compute_dashboard_payload

PAGE_SIZE = 10  # how many rows to show per page on list views


# --------------------------------------------------------------------------- #
# Employees
# --------------------------------------------------------------------------- #

def employee_list(request):
    """Alphabetical employee directory with live search and pagination."""
    qs = Employee.objects.select_related('department').order_by(
        'last_name', 'first_name'
    )
    q = (request.GET.get('q') or '').strip()
    if q:
        qs = qs.filter(
            Q(first_name__icontains=q)
            | Q(last_name__icontains=q)
            | Q(employee_id__icontains=q)
            | Q(position__icontains=q)
            | Q(department__name__icontains=q)
        )

    paginator = Paginator(qs, PAGE_SIZE)
    page_number = request.GET.get('page') or 1
    page_obj = paginator.get_page(page_number)

    querystring = request.GET.copy()
    if 'page' in querystring:
        querystring.pop('page')

    return render(request, 'dtr/employee_list.html', {
        'employees': page_obj.object_list,
        'page_obj': page_obj,
        'paginator': paginator,
        'q': q,
        'querystring': querystring.urlencode(),
    })


def employee_detail(request, pk):
    """Per-employee DTR history (alphabetical by date desc inside the page)."""
    employee = get_object_or_404(Employee, pk=pk)
    records = employee.attendance_records.order_by('-date')
    return render(request, 'dtr/employee_detail.html', {
        'employee': employee,
        'records': records,
    })


@login_required
def employee_create(request):
    if request.method == 'POST':
        form = EmployeeForm(request.POST)
        if form.is_valid():
            emp = form.save()
            messages.success(request, f"Added {emp.full_name}.")
            return redirect('dtr:employee_list')
    else:
        form = EmployeeForm()
    return render(request, 'dtr/employee_form.html', {'form': form, 'mode': 'Add'})


@login_required
def employee_edit(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        form = EmployeeForm(request.POST, instance=employee)
        if form.is_valid():
            form.save()
            messages.success(request, f"Updated {employee.full_name}.")
            return redirect('dtr:employee_list')
    else:
        form = EmployeeForm(instance=employee)
    return render(request, 'dtr/employee_form.html', {
        'form': form, 'mode': 'Edit', 'employee': employee,
    })


# --------------------------------------------------------------------------- #
# Attendance / DTR
# --------------------------------------------------------------------------- #

def attendance_list(request):
    """Master DTR view — alphabetical by employee, then date desc.

    Supports search by employee name/id/department and date filtering.
    Shows PAGE_SIZE records per page with navigation controls.
    """
    qs = Attendance.objects.select_related(
        'employee', 'employee__department'
    ).order_by(
        'employee__last_name', 'employee__first_name', '-date'
    )

    q = (request.GET.get('q') or '').strip()
    if q:
        qs = qs.filter(
            Q(employee__first_name__icontains=q)
            | Q(employee__last_name__icontains=q)
            | Q(employee__employee_id__icontains=q)
            | Q(employee__department__name__icontains=q)
        )

    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)

    # Pagination — 10 per page, preserve other query params.
    paginator = Paginator(qs, PAGE_SIZE)
    page_number = request.GET.get('page') or 1
    page_obj = paginator.get_page(page_number)

    querystring = request.GET.copy()
    if 'page' in querystring:
        querystring.pop('page')

    return render(request, 'dtr/attendance_list.html', {
        'records': page_obj.object_list,
        'page_obj': page_obj,
        'paginator': paginator,
        'q': q,
        'date_from': date_from,
        'date_to': date_to,
        'querystring': querystring.urlencode(),
    })


@require_http_methods(['POST'])
def attendance_update(request, pk):
    """Edit one attendance row in place, then redirect back so the table
    auto-updates with the new value."""
    record = get_object_or_404(Attendance, pk=pk)
    form = AttendanceForm(request.POST, instance=record)
    if form.is_valid():
        form.save()
        messages.success(request, f"Updated record for {record.employee}.")
    else:
        messages.error(request, f"Could not update: {form.errors.as_text()}")
    return redirect(request.META.get('HTTP_REFERER') or 'dtr:attendance_list')


@require_http_methods(['POST'])
def time_in(request, pk):
    """Stamp time-in for an employee for today (or the date in POST)."""
    employee = get_object_or_404(Employee, pk=pk)
    target_date = request.POST.get('date') or date.today().isoformat()
    target_date = date.fromisoformat(target_date)
    now = timezone.now()

    record, _ = Attendance.objects.get_or_create(
        employee=employee, date=target_date
    )
    if not record.time_in:
        record.time_in = now
        record.save(update_fields=['time_in', 'updated_at'])
        messages.success(request, f"Timed in {employee.full_name} at {now:%H:%M}.")
    else:
        messages.info(request, f"{employee.full_name} already timed in at {record.time_in:%H:%M}.")
    return redirect(request.META.get('HTTP_REFERER') or 'dtr:attendance_list')


@require_http_methods(['POST'])
def time_out(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    target_date = request.POST.get('date') or date.today().isoformat()
    target_date = date.fromisoformat(target_date)
    now = timezone.now()

    record = Attendance.objects.filter(employee=employee, date=target_date).first()
    if not record or not record.time_in:
        messages.error(request, f"Cannot time out — {employee.full_name} has no time-in for {target_date}.")
        return redirect(request.META.get('HTTP_REFERER') or 'dtr:attendance_list')

    record.time_out = now
    record.save(update_fields=['time_out', 'updated_at'])
    messages.success(request, f"Timed out {employee.full_name} at {now:%H:%M}.")
    return redirect(request.META.get('HTTP_REFERER') or 'dtr:attendance_list')


# --------------------------------------------------------------------------- #
# Download
# --------------------------------------------------------------------------- #

def _filtered_records(request):
    """Re-use the same filters the list view applies, for downloads."""
    qs = Attendance.objects.select_related(
        'employee', 'employee__department'
    ).order_by('employee__last_name', 'employee__first_name', '-date')
    q = (request.GET.get('q') or '').strip()
    if q:
        qs = qs.filter(
            Q(employee__first_name__icontains=q)
            | Q(employee__last_name__icontains=q)
            | Q(employee__employee_id__icontains=q)
            | Q(employee__department__name__icontains=q)
        )
    if request.GET.get('date_from'):
        qs = qs.filter(date__gte=request.GET['date_from'])
    if request.GET.get('date_to'):
        qs = qs.filter(date__lte=request.GET['date_to'])
    return qs


def download_csv(request):
    qs = _filtered_records(request)
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    filename = f"DTR_{date.today().isoformat()}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow([
        'Employee ID', 'Last Name', 'First Name', 'Department', 'Position',
        'Date', 'Time In', 'Time Out', 'Total Hours', 'Status', 'Remarks',
    ])
    for r in qs:
        writer.writerow([
            r.employee.employee_id,
            r.employee.last_name,
            r.employee.first_name,
            r.employee.department.name if r.employee.department else '',
            r.employee.position,
            r.date.isoformat(),
            r.time_in.strftime('%Y-%m-%d %H:%M') if r.time_in else '',
            r.time_out.strftime('%Y-%m-%d %H:%M') if r.time_out else '',
            r.total_hours if r.total_hours is not None else '',
            r.status,
            r.remarks,
        ])
    return response


def download_excel(request):
    """Excel export using openpyxl. Falls back to a friendly error if missing."""
    try:
        from openpyxl import Workbook
    except ImportError:
        return HttpResponse(
            "openpyxl is not installed. Run: pip install openpyxl",
            status=500,
        )
    qs = _filtered_records(request)
    wb = Workbook()
    ws = wb.active
    ws.title = "DTR"
    headers = [
        'Employee ID', 'Last Name', 'First Name', 'Department', 'Position',
        'Date', 'Time In', 'Time Out', 'Total Hours', 'Status', 'Remarks',
    ]
    ws.append(headers)
    for r in qs:
        # openpyxl cannot serialize tz-aware datetimes — strip tzinfo for Excel.
        ti = r.time_in.replace(tzinfo=None) if r.time_in else ''
        to = r.time_out.replace(tzinfo=None) if r.time_out else ''
        ws.append([
            r.employee.employee_id,
            r.employee.last_name,
            r.employee.first_name,
            r.employee.department.name if r.employee.department else '',
            r.employee.position,
            r.date,
            ti,
            to,
            r.total_hours if r.total_hours is not None else '',
            r.status,
            r.remarks,
        ])
    for col in ws.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(length + 2, 10), 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"DTR_{date.today().isoformat()}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def download_employee_csv(request, pk):
    """Single-employee DTR download."""
    employee = get_object_or_404(Employee, pk=pk)
    qs = employee.attendance_records.order_by('-date')
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    safe = f"{employee.last_name}_{employee.first_name}".replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="DTR_{safe}.csv"'
    writer = csv.writer(response)
    writer.writerow(['Date', 'Time In', 'Time Out', 'Total Hours', 'Status', 'Remarks'])
    for r in qs:
        writer.writerow([
            r.date.isoformat(),
            r.time_in.strftime('%Y-%m-%d %H:%M') if r.time_in else '',
            r.time_out.strftime('%Y-%m-%d %H:%M') if r.time_out else '',
            r.total_hours if r.total_hours is not None else '',
            r.status,
            r.remarks,
        ])
    return response


# --------------------------------------------------------------------------- #
# Dashboard + visualization
# --------------------------------------------------------------------------- #

def dashboard(request):
    """HTML dashboard page. JS will fetch /dashboard/data.json and render charts."""
    return render(request, 'dtr/dashboard.html')


def dashboard_data(request):
    """JSON endpoint used by the dashboard.

    The page polls this every few seconds; if `version` matches the cached
    value it skips re-rendering. When the version differs, charts and
    "Quick Insights" update automatically.
    """
    payload = compute_dashboard_payload()
    return JsonResponse(payload)


@require_http_methods(['POST'])
@login_required
def upload_excel(request):
    """Accept a modified .xlsx (the same format as download_excel),
    re-import rows, then redirect to the dashboard. The new data is
    immediately reflected in the charts on the next poll.
    """
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    f = request.FILES.get('file')
    if not f:
        messages.error(request, 'No file uploaded.')
        return redirect('dtr:dashboard')

    try:
        wb = load_workbook(filename=f, data_only=True)
    except Exception as exc:  # noqa: BLE001
        messages.error(request, f'Could not read the workbook: {exc}')
        return redirect('dtr:dashboard')

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        messages.error(request, 'The workbook is empty.')
        return redirect('dtr:dashboard')

    # Find header row
    header = [str(c).strip() if c is not None else '' for c in rows[0]]
    expected = [
        'Employee ID', 'Last Name', 'First Name', 'Department', 'Position',
        'Date', 'Time In', 'Time Out', 'Total Hours', 'Status', 'Remarks',
    ]
    if [h.lower() for h in header] != [e.lower() for e in expected]:
        messages.error(
            request,
            'Workbook header does not match. Please use the file from '
            '"Download Excel" as the template.',
        )
        return redirect('dtr:dashboard')

    updated = 0
    skipped = 0
    for raw in rows[1:]:
        if not raw or not raw[0]:
            continue
        try:
            emp_id, last, first, dept_name, position, d, ti, to, _, status_v, remarks = raw
        except ValueError:
            skipped += 1
            continue

        emp = Employee.objects.filter(employee_id=emp_id).first()
        if not emp:
            skipped += 1
            continue
        # Update employee fields if they changed in the spreadsheet
        if last and last != emp.last_name:
            emp.last_name = last
        if first and first != emp.first_name:
            emp.first_name = first
        if position and position != emp.position:
            emp.position = position
        if dept_name:
            dept, _ = Department.objects.get_or_create(name=dept_name)
            emp.department = dept
        emp.save()

        if not d:
            continue
        if hasattr(d, 'date'):
            d = d.date()

        rec, _ = Attendance.objects.get_or_create(employee=emp, date=d)
        rec.time_in = ti if hasattr(ti, 'hour') else None
        rec.time_out = to if hasattr(to, 'hour') else None
        rec.remarks = remarks or ''
        rec.save()
        updated += 1

    messages.success(
        request,
        f'Imported {updated} record(s) from Excel. Charts will refresh shortly. (skipped {skipped})',
    )
    return redirect('dtr:dashboard')
